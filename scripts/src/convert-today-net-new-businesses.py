from __future__ import annotations

import hashlib
import ipaddress
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BUNDLE = ROOT / "data/founder-imports/2026-09-06-today-net-new-businesses"
WORKBOOK = BUNDLE / "MWM_TODAY_NET_NEW_BUSINESSES_BUILD022_NYC_LARGE_CUMULATIVE_2026-09-06.xlsx"
OUTPUT = BUNDLE / "today-net-new-business-candidates.jsonl"
SUMMARY = BUNDLE / "conversion-summary.json"
LINK_VALIDATION = BUNDLE / "link-validation.json"
CANONICAL_SHEET = "MANUS_TODAY_NET_NEW"
NYC_SHEET = "BUILD_022_NYC_LARGE"
EXPECTED_SHA256 = "e52dedfc5fb6236ccfd20ace6820c8f63ed6a6f4754597cdee4dc097f04c2dd4"
EXPECTED_ROWS = 3_367
EXPECTED_NYC_ROWS = 170
US_STATE_CODES = set("AL AK AZ AR CA CO CT DE DC FL GA HI ID IL IN IA KS KY LA ME MD MA MI MN MS MO MT NE NV NH NJ NM NY NC ND OH OK OR PA RI SC SD TN TX UT VT VA WA WV WI WY AS GU MP PR VI".split())


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text(value).lower()).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def safe_public_url(value: Any) -> str | None:
    candidate = text(value)
    if not candidate:
        return None
    try:
        parsed = urlparse(candidate)
        if parsed.scheme.lower() not in {"http", "https"} or not parsed.hostname:
            return None
        if parsed.username or parsed.password:
            return None
        host = parsed.hostname.lower().rstrip(".")
        if host == "localhost" or host.endswith(".localhost") or host.endswith(".local"):
            return None
        try:
            ip = ipaddress.ip_address(host)
            if not ip.is_global:
                return None
        except ValueError:
            pass
        return candidate
    except ValueError:
        return None


def social_url(platform: str, value: Any) -> str | None:
    candidate = text(value)
    if not candidate:
        return None
    direct = safe_public_url(candidate)
    if direct:
        host = (urlparse(direct).hostname or "").lower().removeprefix("www.")
        allowed = {
            "instagram": {"instagram.com"},
            "facebook": {"facebook.com", "fb.com"},
            "tiktok": {"tiktok.com"},
        }[platform]
        return direct if any(host == item or host.endswith(f".{item}") for item in allowed) else None
    handle = candidate.lstrip("@").strip("/")
    if not re.fullmatch(r"[A-Za-z0-9._-]{2,80}", handle):
        return None
    if platform == "instagram":
        return f"https://www.instagram.com/{handle}/"
    if platform == "tiktok":
        return f"https://www.tiktok.com/@{handle}"
    return None


def sheet_rows(ws) -> list[dict[str, Any]]:
    iterator = ws.iter_rows(values_only=True)
    headers = [text(value) for value in next(iterator)]
    if not headers or headers[0] != "name":
        raise RuntimeError(f"Unexpected header row in {ws.title}")
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(iterator, start=2):
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        if not any(text(value) for value in row.values()):
            continue
        row["__source_row"] = row_number
        rows.append(row)
    return rows


def identity(row: dict[str, Any]) -> tuple[str, str, str]:
    return norm(row.get("name")), norm(row.get("city")), norm(row.get("state"))


def dedupe_key(row: dict[str, Any]) -> str:
    return f"{norm(row.get('name'))}|{norm(row.get('city'))}|{norm(row.get('state'))}|addr:{norm(row.get('address'))}"


def semicolon_values(value: Any) -> list[str]:
    return sorted({part.strip() for part in re.split(r"[;,]", text(value)) if part.strip()})


def convert(row: dict[str, Any], source_row: int, link_results: dict[str, Any] | None) -> dict[str, Any]:
    name = text(row.get("name"))
    city = text(row.get("city"))
    state = text(row.get("state")).upper()
    address = text(row.get("address"))
    category = text(row.get("category"))
    subcategory = text(row.get("subcategory")) or category
    regulated = text(row.get("regulated_profession")).upper() in {"YES", "REVIEW"}
    recommendation = text(row.get("public_display_recommendation"))
    website = safe_public_url(row.get("website"))
    source_url = safe_public_url(row.get("source_url"))
    social_source_url = safe_public_url(row.get("social_source_url"))
    instagram = social_url("instagram", row.get("instagram_handle"))
    facebook = social_url("facebook", row.get("facebook_url"))
    tiktok = social_url("tiktok", row.get("tiktok_handle"))
    supplied_links = [website, source_url, social_source_url, instagram, facebook, tiktok]
    has_link = any(supplied_links)
    has_reachable_link = has_link if link_results is None else any(
        link and link_results.get(link, {}).get("reachable") is True for link in supplied_links
    )

    gates = ["founder_authorized_cumulative_business", "existing_record_reconciliation"]
    if regulated:
        target_kind = "regulated_review"
        gates.append("regulated_or_licensed_service_evidence_required")
    elif not has_reachable_link:
        target_kind = "manual_review"
        gates.append("reachable_public_link_research_required")
    else:
        target_kind = "business"
        gates.append("ordinary_business_searchable_unclaimed_not_verified")
    gates.append("official_census_address_match_required_before_pin")

    if not name or not city or state not in US_STATE_CODES or not address or not re.search(r"\d", address):
        raise RuntimeError(f"Invalid required business location fields at workbook row {row['__source_row']}")

    ownership_evidence = text(row.get("ownership_or_identity_evidence")) or text(row.get("ownership_group_rollup")) or None
    search_tags = semicolon_values(row.get("need_specific_tags"))
    cultural_specialty = text(row.get("cultural_specialty")) or text(row.get("nearby_cultural_sites_tags")) or None
    return {
        "sourceRow": source_row,
        "sourceWorkbook": WORKBOOK.name,
        "sourceSheet": CANONICAL_SHEET,
        "sourceWorkbookRow": int(row["__source_row"]),
        "targetKind": target_kind,
        "dedupeKey": dedupe_key(row),
        "name": name,
        "city": city,
        "state": state,
        "category": category,
        "subcategory": subcategory,
        "culturalSpecialty": cultural_specialty,
        "address": address,
        "phone": text(row.get("phone")) or None,
        "website": website,
        "sourceUrl": source_url,
        "sourceName": text(row.get("source_name")) or "Founder cumulative business workbook",
        "sourceStatus": text(row.get("source_status")) or None,
        "ownershipDesignations": [],
        "ownershipEvidence": ownership_evidence,
        "regulatedProfession": regulated,
        "publicDisplayRecommendation": recommendation or None,
        "instagramUrl": instagram,
        "facebookUrl": facebook,
        "tiktokUrl": tiktok,
        "socialSourceUrl": social_source_url,
        "priceRange": text(row.get("price_range")) or None,
        "priceBasis": text(row.get("price_basis")) or None,
        "notes": text(row.get("notes")) or None,
        "reviewGates": sorted(set(gates)),
        "rawRecord": {
            "country": "USA",
            "auditDate": text(row.get("audit_date")) or None,
            "locationModel": text(row.get("location_model")) or None,
            "mapPinStatus": text(row.get("map_pin_status")) or None,
            "researchStatus": text(row.get("research_status")) or None,
            "requestedAction": text(row.get("replit_action")) or None,
            "communityConfirmationRule": text(row.get("community_confirmation_rule")) or None,
            "verificationRule": text(row.get("mw_m_verification_rule")) or None,
            "hoursOrAvailability": text(row.get("hours_or_availability")) or None,
            "publicSearchTagEvidence": "workbook_category_services_and_reviewed_offerings_only" if search_tags else None,
            "searchTags": search_tags,
            "originalSourceUrl": text(row.get("source_url")) or None,
            "linkReachabilityAudited": link_results is not None,
        },
    }


def main() -> None:
    actual_sha = sha256(WORKBOOK)
    if actual_sha != EXPECTED_SHA256:
        raise RuntimeError(f"Workbook checksum mismatch: {actual_sha}")
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    canonical_rows = sheet_rows(workbook[CANONICAL_SHEET])
    nyc_rows = sheet_rows(workbook[NYC_SHEET])
    if len(canonical_rows) != EXPECTED_ROWS or len(nyc_rows) != EXPECTED_NYC_ROWS:
        raise RuntimeError(f"Unexpected row counts: canonical={len(canonical_rows)} nyc={len(nyc_rows)}")
    canonical_identities = [identity(row) for row in canonical_rows]
    if len(set(canonical_identities)) != EXPECTED_ROWS:
        raise RuntimeError("Canonical handoff contains duplicate name/city/state identities")
    nyc_identities = {identity(row) for row in nyc_rows}
    if not nyc_identities.issubset(set(canonical_identities)) or len(nyc_identities) != EXPECTED_NYC_ROWS:
        raise RuntimeError("NYC sheet is not an exact identity subset of the canonical cumulative sheet")

    link_results = json.loads(LINK_VALIDATION.read_text(encoding="utf-8")) if LINK_VALIDATION.exists() else None
    candidates = [convert(row, index, link_results) for index, row in enumerate(canonical_rows, start=1)]
    OUTPUT.write_text("".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in candidates), encoding="utf-8")
    target_counts = Counter(row["targetKind"] for row in candidates)
    summary = {
        "workbook": WORKBOOK.name,
        "workbookSha256": actual_sha,
        "canonicalSheet": CANONICAL_SHEET,
        "canonicalRows": len(candidates),
        "nycSubsetRows": len(nyc_rows),
        "targetCounts": dict(sorted(target_counts.items())),
        "ordinaryBusinessRows": target_counts["business"],
        "regulatedReviewRows": target_counts["regulated_review"],
        "missingSafeLinkRows": target_counts["manual_review"],
        "rowsWithStreetAddress": sum(1 for row in candidates if row["address"]),
        "rowsWithWebsite": sum(1 for row in candidates if row["website"]),
        "rowsWithSafeSourceUrl": sum(1 for row in candidates if row["sourceUrl"]),
        "rowsWithSocialLink": sum(1 for row in candidates if any([row["instagramUrl"], row["facebookUrl"], row["tiktokUrl"]])),
        "rowsWithSuppliedCoordinates": 0,
        "manifest": OUTPUT.name,
        "manifestSha256": sha256(OUTPUT),
    }
    SUMMARY.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
