from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BUNDLE = ROOT / "data/founder-imports/2026-09-05-cumulative-content-global"
SOCIAL = BUNDLE / "Mapping-with-Melanin-CUMULATIVE-SOCIAL-NIGHTLIFE-ENRICHMENT-3392.xlsx"
GLOBAL = BUNDLE / "Mapping-with-Melanin-GLOBAL-BUILD-36-CURATED-MINORITY-RECOMMENDED-CUMULATIVE-5086.xlsx"
OUTPUT = BUNDLE / "cumulative-content-global-candidates.jsonl"
SUMMARY = BUNDLE / "conversion-summary.json"

EXPECTED_SOCIAL_SHA = "56163070c7309ed03a2a4fd7344d401cd00e179c9f63a2fef9c798b8ab3840fb"
EXPECTED_GLOBAL_SHA = "db67c6290aa2407b26aabbadebca194c839160ad048dde8a7435f8002ddf5a7c"
EXPECTED_SOCIAL_ROWS = 3392
EXPECTED_DESTINATIONS = 545
EXPECTED_NAMED = 3378
EXPECTED_TOTAL = EXPECTED_SOCIAL_ROWS + EXPECTED_DESTINATIONS + EXPECTED_NAMED

FORCED_RESOURCE_SOURCE_ROWS = {
    23, 49, 55, 60, 67, 118, 215, 259, 465, 2043, 2207, 2252, 2317, 7238, 7302,
    240, 326, 374, 438, 470, 488, 1119, 1122, 1220, 1237, 1484, 1492, 1505, 1767,
    2018, 2020, 2021, 2040, 2042, 2044, 2045, 2052, 2085, 2092, 2128, 2157, 2169,
    2245, 2249, 2354, 2371, 2491, 2585, 2858, 2863, 2868, 2906, 2922, 2949, 3106,
    3107, 3121, 3171, 3235, 3283, 3284, 3291, 3293, 3296, 3297, 3298, 3331, 3337,
    5128, 5130, 5131, 5268, 6280, 6729, 6743, 6806, 6856,
    85, 2156, 2168, 2907, 2936, 3137, 6440, 7274,
    16, 20, 128, 229, 234, 285, 388, 411, 727, 894, 932, 1105, 1120, 1123, 1146,
    1147, 1148, 1236, 1601, 1603, 1663, 1758, 1769, 1770, 2009, 2012, 2041, 2090,
    2158, 2183, 2250, 2287, 2290, 2292, 2315, 2316, 2352, 2353, 2372, 2401, 2567,
    2568, 2711, 2837, 2860, 2861, 2862, 2864, 2905, 3234, 3282, 3292, 4724, 4733,
    4884, 5062, 5129, 5133, 5308, 5403, 5704, 5743, 6301, 6302, 6308, 6416, 6426,
    6728, 6730, 6802,
}
FORCED_MANUAL_SOURCE_ROWS = {
    321, 1138, 1871, 2116, 2544, 3294, 4384, 5742, 5844, 5866, 5867, 6239, 6263,
    6450, 6455, 6798, 6854,
    693, 694, 695, 696, 697, 698, 719, 720, 721, 722, 723, 724,
    1232, 1419, 1761, 1855, 1882, 1973, 2088, 2115, 2127, 2163, 2165, 2206, 2247,
    2312, 2402, 2569, 2570, 2699, 2953, 3075, 3095, 3133, 3295, 3299, 4747, 5496,
    5678, 5831, 5832, 5837, 5839, 5860, 5861, 5862, 5863, 5864, 5868, 5869, 5871,
    5872, 6411, 6727, 6774,
    1988, 2755, 2833, 2950, 2981, 5739, 6794,
    439, 442, 474, 475, 1243, 1280, 1535, 2010, 2033, 2129, 2130, 2160, 2166,
    2267, 2273, 2550, 2749, 5092, 5093, 5358, 7306,
}

HEADER_TERMS = {
    "city", "state", "name", "category", "destination", "country/territory",
    "latitude", "longitude", "master status", "mwm map status", "website",
    "primary source url", "safety source", "research date", "id",
}
RESOURCE_RE = re.compile(
    r"\b(resources?|programs?|grants?|assistance|nonprofits?|non-profits?|foundations?|museums?|heritage|"
    r"community organizations?|community cent(?:er|re)s?|cultural organizations?|arts?\s*/\s*cultural organizations?|"
    r"business organizations?|chambers? of commerce|housing authorit(?:y|ies)|farmworker cent(?:er|re)s?|"
    r"forums? for arts|municipal\s*/\s*diaspora networks?|business\s*/\s*community networks?|"
    r"\bcdc\b|offices? of business opportunities|public health departments?|head start|ministr(?:y|ies)|"
    r"advisory councils?|civil rights institutes?|support teams?|black chambers?|"
    r"coalitions?|associations?|coordinated entry systems?|development corporations?|"
    r"homelessness systems?|homeless resource systems?|resource cent(?:er|re)s?|workforce|training|"
    r"senior services?|housing counseling|caregiver support|school|college|university|health hub|"
    r"government|public service|shelter|housing help|food bank|library)\b",
    re.I,
)
NON_BUSINESS_PLACE_RE = re.compile(
    r"\b(festivals?|recurring(?:\s+[a-z&/-]+){0,3}\s+(?:festivals?|fairs?|conferences?|cultural gatherings?|part(?:y|ies)|markets?)|"
    r"events?|carnivals?|parades?|neighbou?rhoods?|districts?|locations?|waterfronts?|boardwalks?|caves?|"
    r"national parks?|public parks?|parks?|beach(?:es)?|landmarks?|attractions?|"
    r"historic(?:al)?\s*(?:/\s*religious\s*)?sites?|historic communities?|historic districts?|historic cemeter(?:y|ies)|"
    r"sacred\s*/\s*(?:historic|cultural)|religious sites?|heritage sites?|cultural sites?|"
    r"public spaces?|public art|murals?|monuments?|memorials?|cemeter(?:y|ies)|"
    r"trails?|squares?|plazas?|waterfalls?|islands?|museums?|cultural cent(?:er|re)s?|"
    r"temples?|church(?:es)?|cathedrals?|basilicas?|synagogues?|gurdwaras?|mosques?|shrines?|"
    r"palaces?|castles?|forts?|gardens?|zoos?|aquariums?)\b",
    re.I,
)
STRONG_NAMED_PLACE_RE = re.compile(
    r"\b(national park|historic(?:al)? site|historic cemetery|historic district|historic community|"
    r"historic(?:al)? [a-z&'-]+ (?:beach )?park|historic market square|religious site|heritage site|sacred site|"
    r"public park|museum|memorial|monument|festival|carnival|parade|public beach|public garden|"
    r"national performing arts venue|historic institution|historic house|civil rights site|historic hotel site|"
    r"non-commerce cultural pin|old market stalls|business expo|\w+ fest|market popups|cultural marketplace|"
    r"boardwalk|waterfront|cemetery|cathedral|basilica|synagogue|gurdwara|mosque|church)\b",
    re.I,
)
REGULATED_RE = re.compile(
    r"\b(doctor|physician|medical|clinic|therap|dentist|pharmacy|attorney|lawyer|legal|"
    r"accountant|cpa|financial advisor|insurance|mortgage|realtor|real estate|childcare|"
    r"daycare|preschool|cannabis|dispensary|electrician|plumber|hvac|roofing|contractor)\b",
    re.I,
)


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text(value).lower()).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def public_url(value: Any) -> str | None:
    candidate = text(value)
    return candidate if re.match(r"^https?://", candidate, re.I) else None


def find_header(ws) -> tuple[int | None, list[str]]:
    best_row: int | None = None
    best_headers: list[str] = []
    best_score = 0
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 15, 15), values_only=True), start=1):
        headers = [text(value) for value in row]
        score = sum(1 for value in headers if value.lower() in HEADER_TERMS)
        if score > best_score:
            best_score = score
            best_row = row_num
            best_headers = headers
    return (best_row, best_headers) if best_score >= 2 else (None, [])


def rows_for_sheet(ws) -> tuple[list[str], list[dict[str, Any]], int | None]:
    header_row, headers = find_header(ws)
    if header_row is None:
        return [], [], None
    unique_headers: list[str] = []
    seen: Counter[str] = Counter()
    for index, header in enumerate(headers):
        key = header or f"column_{index + 1}"
        seen[key] += 1
        unique_headers.append(key if seen[key] == 1 else f"{key}_{seen[key]}")
    rows: list[dict[str, Any]] = []
    for source_row, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(text(value) for value in values):
            continue
        row = {unique_headers[i]: values[i] if i < len(values) else None for i in range(len(unique_headers))}
        row["__source_row"] = source_row
        rows.append(row)
    return unique_headers, rows, header_row


def dedupe_key(name: str, city: str, state: str, address: str | None = None) -> str:
    base = f"{norm(name)}|{norm(city)}|{norm(state)}"
    return f"{base}|addr:{norm(address)}" if address else f"{base}|no-location"


def existing_candidate_keys() -> set[tuple[str, str, str]]:
    keys: set[tuple[str, str, str]] = set()
    for path in [
        ROOT / "data/founder-imports/2026-09-04/directory-import-candidates.jsonl",
        ROOT / "data/founder-imports/2026-09-05-kinfolk-poc-businesses/kinfolk-poc-business-candidates.jsonl",
    ]:
        if not path.exists():
            continue
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                row = json.loads(line)
                keys.add((norm(row.get("name")), norm(row.get("city")), norm(row.get("state"))))
    return keys


def classify_identity(name: str, category: str, description: str) -> tuple[str, list[str]]:
    identity = f"{name} {category} {description}"
    if norm(name) == "date" or re.fullmatch(r"20\d{2}[-/]\d{2}[-/]\d{2}", text(name)):
        return "manual_review", ["source_metadata_artifact"]
    if re.search(r"\btours?(?:\s*/\s*experience)?\b", category, re.I) and re.search(r"\b(tour|experience)\b", name, re.I):
        return "manual_review", ["tour_operator_identity_required"]
    if NON_BUSINESS_PLACE_RE.search(f"{category} {description}") or STRONG_NAMED_PLACE_RE.search(f"{name} {description}"):
        return "manual_review", ["event_or_place_destination_required"]
    if RESOURCE_RE.search(identity):
        return "community_resource", ["business_vs_resource_routing_required"]
    if REGULATED_RE.search(identity):
        return "regulated_review", ["regulated_or_licensed_service_evidence_required"]
    return "business", []


def social_candidates(start: int, prior: set[tuple[str, str, str]]) -> list[dict[str, Any]]:
    wb = load_workbook(SOCIAL, read_only=True, data_only=True)
    ws = wb["MWM Content Bible"]
    _headers, rows, _header_row = rows_for_sheet(ws)
    meaningful = [row for row in rows if text(row.get("Name")) and text(row.get("City"))]
    identity_counts = Counter((norm(row.get("Name")), norm(row.get("City")), norm(row.get("State"))) for row in meaningful)
    candidates: list[dict[str, Any]] = []
    for offset, row in enumerate(meaningful):
        name = text(row.get("Name"))
        city = text(row.get("City"))
        state = text(row.get("State"))
        category = text(row.get("Category")) or "Manual classification required"
        story_type = text(row.get("Story Type"))
        key = (norm(name), norm(city), norm(state))
        target, gates = classify_identity(name, category, story_type)
        gates.extend(["street_address_required_before_publication", "precise_geocode_required_before_pin"])
        if identity_counts[key] > 1:
            gates.append("duplicate_within_workbook")
        if key in prior:
            gates.append("prior_founder_candidate_reconciliation")
        website = public_url(row.get("Website"))
        source_url = public_url(row.get("Primary Source URL"))
        instagram = public_url(row.get("Instagram URL"))
        tiktok = public_url(row.get("TikTok URL"))
        facebook = public_url(row.get("Facebook"))
        if not any([website, source_url, instagram, tiktok, facebook]):
            gates.append("public_link_required_before_publication")
        else:
            gates.append("public_links_require_current_validation")
        candidates.append({
            "sourceRow": start + offset,
            "sourceWorkbook": SOCIAL.name,
            "sourceSheet": ws.title,
            "sourceWorkbookRow": row["__source_row"],
            "targetKind": target,
            "dedupeKey": dedupe_key(name, city, state),
            "name": name,
            "city": city,
            "state": state,
            "category": category,
            "subcategory": story_type or None,
            "culturalSpecialty": text(row.get("City-Specific Tag(s)")) or None,
            "address": None,
            "website": website,
            "sourceUrl": source_url,
            "sourceName": "Founder cumulative social/nightlife workbook",
            "sourceStatus": text(row.get("Master Status")) or None,
            "ownershipDesignations": [],
            "ownershipEvidence": story_type or None,
            "regulatedProfession": target == "regulated_review",
            "publicDisplayRecommendation": "Hold until destination, operating status, exact street location, and links are confirmed.",
            "instagramUrl": instagram,
            "facebookUrl": facebook,
            "tiktokUrl": tiktok,
            "socialSourceUrl": public_url(row.get("Social Verification Source")),
            "notes": text(row.get("Required QA")) or text(row.get("Legacy Social Note")) or None,
            "reviewGates": sorted(set(gates)),
            "rawRecord": {
                "cityCode": text(row.get("City Code")) or None,
                "storyType": story_type or None,
                "whyFeature": text(row.get("Why MWM Should Feature It")) or None,
                "contentAngle": text(row.get("Content Angle")) or None,
                "cityRelevance": text(row.get("City Relevance")) or None,
                "masterGapFlag": text(row.get("Master Gap Flag")) or None,
                "socialResearchStatus": text(row.get("Social Research Status")) or None,
                "socialVerificationDate": text(row.get("Social Verification Date")) or None,
            },
        })
    wb.close()
    if len(candidates) != EXPECTED_SOCIAL_ROWS:
        raise RuntimeError(f"Expected {EXPECTED_SOCIAL_ROWS} social rows, received {len(candidates)}")
    return candidates


def global_candidates(start: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(GLOBAL, read_only=True, data_only=True)
    destination_unique: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    named_unique: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for ws in wb.worksheets:
        headers, rows, _header_row = rows_for_sheet(ws)
        destination_field = next((field for field in ["Destination", "Place / Destination"] if field in headers), None)
        name_field = next((field for field in ["Name", "Business / Named Entity"] if field in headers), None)
        if destination_field:
            for row in rows:
                destination = text(row.get(destination_field))
                country = text(row.get("Country/Territory")) or text(row.get("Country"))
                latitude = text(row.get("Latitude"))
                longitude = text(row.get("Longitude"))
                if not destination:
                    continue
                key = (norm(destination), norm(country), latitude, longitude)
                destination_unique.setdefault(key, {"sheet": ws.title, "row": row})
        if name_field:
            for row in rows:
                name = text(row.get(name_field))
                country = text(row.get("Country/Territory")) or text(row.get("Country"))
                city_area = text(row.get("City / Area"))
                pin_type = text(row.get("Pin Type")) or text(row.get("Map Point Type")) or text(row.get("Business Type"))
                if not name:
                    continue
                key = (norm(name), norm(country), norm(city_area), norm(pin_type))
                named_unique.setdefault(key, {"sheet": ws.title, "row": row, "name": name, "country": country, "cityArea": city_area, "pinType": pin_type})

    destinations: list[dict[str, Any]] = []
    for offset, item in enumerate(destination_unique.values()):
        row = item["row"]
        name = text(row.get("Destination")) or text(row.get("Place / Destination"))
        country = text(row.get("Country/Territory")) or text(row.get("Country"))
        region = text(row.get("Region"))
        primary_source = public_url(row.get("Primary Travel Source")) or public_url(row.get("Primary Research Source"))
        qa_note = text(row.get("Map QA Note"))
        gates = ["travel_destination_layer_required", "destination_coordinates_not_business_coordinates"]
        if "approximate" in qa_note.lower() or "centroid" in qa_note.lower():
            gates.append("approximate_or_centroid_coordinate_review")
        if not primary_source:
            gates.append("primary_travel_source_required")
        destinations.append({
            "sourceRow": start + offset,
            "sourceWorkbook": GLOBAL.name,
            "sourceSheet": item["sheet"],
            "sourceWorkbookRow": row["__source_row"],
            "targetKind": "manual_review",
            "dedupeKey": dedupe_key(name, name, country),
            "name": name,
            "city": name,
            "state": region or country or "International",
            "category": "Travel Destination",
            "subcategory": text(row.get("Map Point Type")) or None,
            "address": None,
            "website": None,
            "sourceUrl": primary_source,
            "sourceName": "Founder global curated destination workbook",
            "sourceStatus": text(row.get("MWM Map Status")) or None,
            "ownershipDesignations": [],
            "ownershipEvidence": None,
            "regulatedProfession": False,
            "publicDisplayRecommendation": "Route to the travel destination layer; never publish as a business pin.",
            "notes": qa_note or None,
            "reviewGates": sorted(set(gates)),
            "rawRecord": {
                "destination": name,
                "country": country,
                "region": region or None,
                "latitude": text(row.get("Latitude")) or None,
                "longitude": text(row.get("Longitude")) or None,
                "geocodeQuery": text(row.get("Geocode Query")) or None,
                "evidenceScope": text(row.get("Evidence Scope")) or None,
                "evidencePrecision": text(row.get("Evidence Precision")) or None,
                "safetySource": public_url(row.get("Safety Source")),
                "safetyNotes": text(row.get("Safety Notes")) or None,
                "researchDate": text(row.get("Research Date")) or None,
            },
        })

    named: list[dict[str, Any]] = []
    for offset, item in enumerate(named_unique.values()):
        row = item["row"]
        name = item["name"]
        country = item["country"]
        city_area = item["cityArea"]
        pin_type = item["pinType"] or text(row.get("Category")) or "Named place"
        description = text(row.get("Why It Belongs"))
        secondary_category = text(row.get("Category"))
        target, gates = classify_identity(name, pin_type, f"{secondary_category} {description}")
        gates.extend([
            "country_or_area_only_location",
            "street_address_required_before_publication",
            "precise_geocode_required_before_pin",
            "current_operating_status_required",
            "official_business_link_required",
        ])
        source_url = public_url(row.get("Primary Source")) or public_url(row.get("Primary Research Source"))
        if source_url:
            gates.append("third_party_source_requires_current_validation")
        if re.search(r"owned|minority|black|woman|women|latino|latina|indigenous", pin_type, re.I):
            gates.append("ownership_evidence_review")
        city = city_area or country or "Location research required"
        state = country or "International"
        named.append({
            "sourceRow": start + len(destinations) + offset,
            "sourceWorkbook": GLOBAL.name,
            "sourceSheet": item["sheet"],
            "sourceWorkbookRow": row["__source_row"],
            "targetKind": target,
            "dedupeKey": dedupe_key(name, city, state),
            "name": name,
            "city": city,
            "state": state,
            "category": pin_type,
            "subcategory": secondary_category or None,
            "address": None,
            "website": None,
            "sourceUrl": source_url,
            "sourceName": "Founder global curated named-entity workbook",
            "sourceStatus": text(row.get("MWM Status")) or text(row.get("Verification Stage")) or None,
            "ownershipDesignations": [],
            "ownershipEvidence": pin_type if "owned" in pin_type.lower() else None,
            "regulatedProfession": target == "regulated_review",
            "publicDisplayRecommendation": "Hold until exact destination, official link, current operating status, and precise location are confirmed.",
            "priceRange": text(row.get("Budget Tier")) or None,
            "priceBasis": text(row.get("Budget Class")) or None,
            "notes": text(row.get("Required QA")) or text(row.get("Map QA Note")) or None,
            "reviewGates": sorted(set(gates)),
            "rawRecord": {
                "cumulativeId": text(row.get("Cumulative ID")) or text(row.get("Clean Pin ID")) or None,
                "country": country,
                "cityArea": city_area or None,
                "pinType": pin_type,
                "whyItBelongs": description or None,
                "travelerValue": text(row.get("MWM / Black Traveler Value")) or None,
                "safetySource": public_url(row.get("Safety Source")),
                "researchDate": text(row.get("Research Date")) or None,
            },
        })
    wb.close()
    if len(destinations) != EXPECTED_DESTINATIONS:
        raise RuntimeError(f"Expected {EXPECTED_DESTINATIONS} destinations, received {len(destinations)}")
    if len(named) != EXPECTED_NAMED:
        raise RuntimeError(f"Expected {EXPECTED_NAMED} named entities, received {len(named)}")
    return destinations, named


def main() -> None:
    if sha256(SOCIAL) != EXPECTED_SOCIAL_SHA or sha256(GLOBAL) != EXPECTED_GLOBAL_SHA:
        raise RuntimeError("Workbook checksum mismatch; refusing conversion")
    prior = existing_candidate_keys()
    social = social_candidates(1, prior)
    destinations, named = global_candidates(len(social) + 1)
    candidates = social + destinations + named
    if len(candidates) != EXPECTED_TOTAL:
        raise RuntimeError(f"Expected {EXPECTED_TOTAL} candidates, received {len(candidates)}")
    if len({row["sourceRow"] for row in candidates}) != EXPECTED_TOTAL:
        raise RuntimeError("Synthetic source rows are not unique")
    for row in candidates:
        source_row = row["sourceRow"]
        if source_row in FORCED_RESOURCE_SOURCE_ROWS:
            row["targetKind"] = "community_resource"
            row["reviewGates"].append("reviewed_resource_identity_routing")
        if source_row in FORCED_MANUAL_SOURCE_ROWS:
            row["targetKind"] = "manual_review"
            row["reviewGates"].append("reviewed_non_business_or_metadata_routing")
    if any(
        row["targetKind"] != "community_resource"
        for row in candidates
        if row["sourceRow"] in FORCED_RESOURCE_SOURCE_ROWS
    ):
        raise RuntimeError("Every reviewed resource identity must route to community_resource")
    if any(
        row["targetKind"] != "manual_review"
        for row in candidates
        if row["sourceRow"] in FORCED_MANUAL_SOURCE_ROWS
    ):
        raise RuntimeError("Every reviewed event, place, tour, or metadata artifact must route to manual_review")
    final_key_counts = Counter(row["dedupeKey"] for row in candidates)
    final_key_targets: dict[str, set[str]] = defaultdict(set)
    for row in candidates:
        final_key_targets[row["dedupeKey"]].add(row["targetKind"])
    for row in candidates:
        if final_key_counts[row["dedupeKey"]] > 1:
            row["reviewGates"].append("duplicate_within_workbook")
        if len(final_key_targets[row["dedupeKey"]]) > 1:
            row["reviewGates"].append("conflicting_target_routing")
        row["reviewGates"] = sorted(set(row["reviewGates"]))
    if any(
        "duplicate_within_workbook" not in row["reviewGates"]
        for row in candidates
        if final_key_counts[row["dedupeKey"]] > 1
    ):
        raise RuntimeError("Every repeated final dedupe key must be gated")
    if any(
        row["targetKind"] != "manual_review"
        for row in candidates
        if NON_BUSINESS_PLACE_RE.search(row["category"])
        or STRONG_NAMED_PLACE_RE.search(row["name"])
    ):
        raise RuntimeError("Every explicit event or place category must route to manual review")
    resource_business_misroutes = [
        row["sourceRow"]
        for row in candidates
        if row["targetKind"] == "business"
        and RESOURCE_RE.search(
            " ".join([
                text(row.get("name")),
                text(row.get("category")),
                text(row.get("subcategory")),
                text((row.get("rawRecord") or {}).get("whyItBelongs")),
            ])
        )
    ]
    if resource_business_misroutes:
        raise RuntimeError(
            f"Resource identities cannot route to business: {resource_business_misroutes[:20]}"
        )
    public_place_business_misroutes = [
        row["sourceRow"]
        for row in candidates
        if row["targetKind"] == "business"
        and (
            NON_BUSINESS_PLACE_RE.search(
                " ".join([
                    text(row.get("category")),
                    text(row.get("subcategory")),
                    text((row.get("rawRecord") or {}).get("whyItBelongs")),
                ])
            )
            or STRONG_NAMED_PLACE_RE.search(text(row.get("name")))
        )
    ]
    if public_place_business_misroutes:
        raise RuntimeError(
            f"Event and public-place identities cannot route to business: {public_place_business_misroutes[:20]}"
        )
    with OUTPUT.open("w", encoding="utf-8", newline="\n") as handle:
        for row in candidates:
            handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
    target_counts = Counter(row["targetKind"] for row in candidates)
    gate_counts: Counter[str] = Counter(gate for row in candidates for gate in row["reviewGates"])
    summary = {
        "sourceChecksums": {SOCIAL.name: EXPECTED_SOCIAL_SHA, GLOBAL.name: EXPECTED_GLOBAL_SHA},
        "candidateRows": len(candidates),
        "socialRows": len(social),
        "travelDestinationRows": len(destinations),
        "globalNamedEntityRows": len(named),
        "targetCounts": dict(sorted(target_counts.items())),
        "reviewGateCounts": dict(sorted(gate_counts.items())),
        "manifestSha256": sha256(OUTPUT),
        "publicationWrites": 0,
    }
    SUMMARY.write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise
